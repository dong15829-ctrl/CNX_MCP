from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

from sqlalchemy import func, select, update, delete
from sqlalchemy.orm import Session

from di_es_dashboard_api.models import (
    CountryMetric,
    CountrySummary,
    IssueAltText,
    IssueBottom30,
    ReportFile,
    ReportIngestLog,
)
from di_es_dashboard_api.services.excel_parser import parse_alt_text_issues, parse_country_table_from_sheet
from di_es_dashboard_api.utils import parse_report_filename, sha256_file


def _log(db: Session, report_file_id: int, level: str, message: str, context: dict[str, Any] | None = None) -> None:
    db.add(ReportIngestLog(report_file_id=report_file_id, level=level, message=message, context=context))


def next_revision(db: Session, report_type: str, year: int, month: int) -> int:
    value = db.execute(
        select(func.max(ReportFile.revision)).where(
            ReportFile.report_type == report_type, ReportFile.year == year, ReportFile.month == month
        )
    ).scalar_one_or_none()
    return int(value or 0) + 1


def deactivate_completed_active(db: Session, report_type: str, year: int, month: int) -> None:
    db.execute(
        update(ReportFile)
        .where(
            ReportFile.report_type == report_type,
            ReportFile.year == year,
            ReportFile.month == month,
            ReportFile.is_active.is_(True),
        )
        .values(is_active=False)
    )


def purge_report_data(db: Session, report_file_id: int) -> None:
    db.execute(delete(CountryMetric).where(CountryMetric.report_file_id == report_file_id))
    db.execute(delete(CountrySummary).where(CountrySummary.report_file_id == report_file_id))
    db.execute(delete(IssueAltText).where(IssueAltText.report_file_id == report_file_id))
    db.execute(delete(IssueBottom30).where(IssueBottom30.report_file_id == report_file_id))


def ingest_excel_report(db: Session, report_file: ReportFile) -> None:
    file_path = Path(report_file.stored_path)
    report_type, year, month = report_file.report_type, report_file.year, report_file.month

    report_file.sha256 = sha256_file(file_path)
    report_file.status = "processing"
    report_file.error_message = None
    report_file.warnings_count = 0
    db.add(report_file)
    db.flush()

    purge_report_data(db, report_file.id)
    _log(db, report_file.id, "INFO", "Starting ingestion", {"report_type": report_type, "year": year, "month": month})

    if report_type == "B2B":
        summary_sheet = "Summary (최종)" if "Summary (최종)" in _sheetnames(file_path) else "Summary"
    elif report_type == "B2C":
        summary_sheet = "Summary by Country"
    else:
        raise ValueError(f"Unsupported report_type: {report_type}")

    try:
        summary_rows = parse_country_table_from_sheet(file_path, summary_sheet)
    except Exception as e:
        _log(db, report_file.id, "ERROR", "Failed to parse summary sheet", {"sheet": summary_sheet, "error": str(e)})
        raise

    country_summaries: list[CountrySummary] = []
    country_metrics: list[CountryMetric] = []

    for r in summary_rows:
        country_summaries.append(
            CountrySummary(
                report_file_id=report_file.id,
                region=r.region,
                country=r.country,
                gp1_status=r.gp1_status,
                sku=r.sku,
                sku_prev=r.sku_prev,
                sku_gap=r.sku_gap,
                total_score=r.total_score,
                total_score_prev=r.total_score_prev,
                total_gap=r.total_gap,
                rank_prev=r.rank_prev,
                rank_current=r.rank_current,
                rank_change=r.rank_change,
            )
        )

        for metric_key, (metric_label, value) in r.metrics.items():
            country_metrics.append(
                CountryMetric(
                    report_file_id=report_file.id,
                    country=r.country,
                    metric_key=metric_key,
                    metric_label=metric_label,
                    group_label=None,
                    value=value,
                )
            )

    db.bulk_save_objects(country_summaries)
    db.bulk_save_objects(country_metrics)
    _log(
        db,
        report_file.id,
        "INFO",
        "Summary imported",
        {"countries": len(country_summaries), "metrics": len(country_metrics), "sheet": summary_sheet},
    )

    issues_alt = parse_alt_text_issues(file_path)
    if issues_alt:
        db.bulk_save_objects(
            [
                IssueAltText(
                    report_file_id=report_file.id,
                    country=i.country,
                    url=i.url,
                    card_index=i.card_index,
                    src=i.src,
                    srcset=i.srcset,
                    alt=i.alt,
                    alt_length=i.alt_length,
                    alt_comment=i.alt_comment,
                )
                for i in issues_alt
            ]
        )
        _log(db, report_file.id, "INFO", "Alt text issues imported", {"count": len(issues_alt)})

    for bottom_sheet in ["Bottom 30%"]:
        if bottom_sheet not in _sheetnames(file_path):
            continue
        try:
            bottom_rows = parse_country_table_from_sheet(file_path, bottom_sheet)
        except Exception as e:
            _log(db, report_file.id, "WARN", "Failed to parse bottom30 sheet", {"sheet": bottom_sheet, "error": str(e)})
            continue

        issues_bottom = [
            IssueBottom30(
                report_file_id=report_file.id,
                region=r.region,
                country=r.country,
                total_score=r.total_score,
                rank_prev=r.rank_prev,
                rank_current=r.rank_current,
                rank_change=r.rank_change,
                details={
                    "sku": r.sku,
                    "sku_prev": r.sku_prev,
                    "sku_gap": r.sku_gap,
                    "total_score_prev": r.total_score_prev,
                    "total_gap": r.total_gap,
                    "metrics": {k: v for k, (_, v) in r.metrics.items()},
                },
            )
            for r in bottom_rows
        ]
        if issues_bottom:
            db.bulk_save_objects(issues_bottom)
            _log(db, report_file.id, "INFO", "Bottom30 imported", {"count": len(issues_bottom), "sheet": bottom_sheet})

    report_file.status = "completed"
    report_file.ingested_at = dt.datetime.utcnow()
    db.add(report_file)

    deactivate_completed_active(db, report_type, year, month)
    report_file.is_active = True
    db.add(report_file)
    _log(db, report_file.id, "INFO", "Ingestion completed")


def _sheetnames(file_path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def create_report_file_record(db: Session, stored_path: Path, original_name: str) -> ReportFile:
    report_type, year, month = parse_report_filename(original_name)
    revision = next_revision(db, report_type, year, month)
    report_file = ReportFile(
        report_type=report_type,
        year=year,
        month=month,
        revision=revision,
        file_name=original_name,
        stored_path=str(stored_path),
        status="pending",
        is_active=False,
    )
    db.add(report_file)
    db.flush()
    return report_file

