from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from di_es_dashboard_api.utils import normalize_header, slugify_metric, to_float, to_int, clean_cell


@dataclass(frozen=True)
class ParsedCountryRow:
    region: str | None
    country: str
    gp1_status: str | None
    sku: int | None
    sku_prev: int | None
    sku_gap: int | None
    total_score: float | None
    total_score_prev: float | None
    total_gap: float | None
    rank_prev: int | None
    rank_current: int | None
    rank_change: int | None
    metrics: dict[str, tuple[str, float | None]]


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 30) -> int:
    required = {"REGION", "COUNTRY"}
    for r in range(1, max_rows + 1):
        row = [normalize_header(c.value) for c in ws[r]]
        if required.issubset(set(row)):
            return r
    raise ValueError("Header row not found (REGION/COUNTRY)")


def _build_base_col_index(header_values: list[str]) -> dict[str, int]:
    index_by_value: dict[str, list[int]] = {}
    for i, h in enumerate(header_values):
        if not h:
            continue
        index_by_value.setdefault(h, []).append(i)

    region_i = index_by_value.get("REGION", [None])[0]
    country_i = index_by_value.get("COUNTRY", [None])[0]
    if region_i is None or country_i is None:
        raise ValueError("Missing REGION/COUNTRY columns")

    gp1_i = next((i for i, h in enumerate(header_values) if "GP1" in h and "STATUS" in h), None)

    sku_i = next((i for i, h in enumerate(header_values) if h == "SKU"), None)
    if sku_i is None:
        sku_i = next((i for i, h in enumerate(header_values) if h.startswith("SKU")), None)

    sku_prev_i = None
    if sku_i is not None:
        sku_prev_i = next(
            (i for i in range(sku_i + 1, len(header_values)) if header_values[i].startswith("SKU") and header_values[i] != "SKU"),
            None,
        )

    sku_gap_i = None
    if sku_prev_i is not None:
        sku_gap_i = next((i for i in range(sku_prev_i + 1, len(header_values)) if header_values[i] == "GAP"), None)

    total_score_idxs = [i for i, h in enumerate(header_values) if "TOTAL" in h and "SCORE" in h]
    total_i = total_score_idxs[0] if len(total_score_idxs) >= 1 else None
    total_prev_i = total_score_idxs[1] if len(total_score_idxs) >= 2 else None

    total_gap_i = None
    if total_prev_i is not None:
        total_gap_i = next((i for i in range(total_prev_i + 1, len(header_values)) if header_values[i] == "GAP"), None)

    return {
        "region": region_i,
        "country": country_i,
        "gp1_status": gp1_i if gp1_i is not None else -1,
        "sku": sku_i if sku_i is not None else -1,
        "sku_prev": sku_prev_i if sku_prev_i is not None else -1,
        "sku_gap": sku_gap_i if sku_gap_i is not None else -1,
        "total_score": total_i if total_i is not None else -1,
        "total_score_prev": total_prev_i if total_prev_i is not None else -1,
        "total_gap": total_gap_i if total_gap_i is not None else -1,
    }


def _build_metric_col_index(
    ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int
) -> tuple[dict[int, tuple[str, str]], dict[str, int]]:
    metric_row_idx = header_row + 2
    metric_row = [clean_cell(c.value) for c in ws[metric_row_idx]]

    metric_cols: dict[int, tuple[str, str]] = {}
    rank_cols: dict[str, int] = {}

    for col_idx, raw in enumerate(metric_row):
        if raw is None:
            continue
        label = str(raw).strip()
        if not label:
            continue

        upper = normalize_header(label)
        if "RANK" in upper:
            if "OCT" in upper or "OCTOBER" in upper:
                rank_cols["rank_prev"] = col_idx
            elif "NOV" in upper or "NOVEMBER" in upper:
                rank_cols["rank_current"] = col_idx
            else:
                rank_cols["rank_current"] = col_idx
            continue
        if upper == "CHANGE":
            rank_cols["rank_change"] = col_idx
            continue

        key = slugify_metric(label)
        metric_cols[col_idx] = (key, label)

    return metric_cols, rank_cols


def parse_country_table_from_sheet(file_path: Path, sheet_name: str) -> list[ParsedCountryRow]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        header_values = [normalize_header(c.value) for c in ws[header_row]]
        base_idx = _build_base_col_index(header_values)
        metric_cols, rank_cols = _build_metric_col_index(ws, header_row)

        data_start = header_row + 5
        parsed: list[ParsedCountryRow] = []

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            country_raw = row[base_idx["country"]] if base_idx["country"] >= 0 else None
            country = clean_cell(country_raw)
            if not country:
                continue
            if str(country).strip() == "*":
                continue

            region = clean_cell(row[base_idx["region"]]) if base_idx["region"] >= 0 else None
            gp1_status = None
            if base_idx["gp1_status"] >= 0:
                gp1_status = clean_cell(row[base_idx["gp1_status"]])

            sku = to_int(row[base_idx["sku"]]) if base_idx["sku"] >= 0 else None
            sku_prev = to_int(row[base_idx["sku_prev"]]) if base_idx["sku_prev"] >= 0 else None
            sku_gap = to_int(row[base_idx["sku_gap"]]) if base_idx["sku_gap"] >= 0 else None

            total_score = to_float(row[base_idx["total_score"]]) if base_idx["total_score"] >= 0 else None
            total_score_prev = to_float(row[base_idx["total_score_prev"]]) if base_idx["total_score_prev"] >= 0 else None
            total_gap = to_float(row[base_idx["total_gap"]]) if base_idx["total_gap"] >= 0 else None

            rank_prev = to_int(row[rank_cols["rank_prev"]]) if "rank_prev" in rank_cols else None
            rank_current = to_int(row[rank_cols["rank_current"]]) if "rank_current" in rank_cols else None
            rank_change = to_int(row[rank_cols["rank_change"]]) if "rank_change" in rank_cols else None

            metrics: dict[str, tuple[str, float | None]] = {}
            for col_idx, (m_key, m_label) in metric_cols.items():
                if col_idx >= len(row):
                    continue
                metrics[m_key] = (m_label, to_float(row[col_idx]))

            parsed.append(
                ParsedCountryRow(
                    region=str(region) if region is not None else None,
                    country=str(country),
                    gp1_status=str(gp1_status) if gp1_status is not None else None,
                    sku=sku,
                    sku_prev=sku_prev,
                    sku_gap=sku_gap,
                    total_score=total_score,
                    total_score_prev=total_score_prev,
                    total_gap=total_gap,
                    rank_prev=rank_prev,
                    rank_current=rank_current,
                    rank_change=rank_change,
                    metrics=metrics,
                )
            )

        return parsed
    finally:
        wb.close()


@dataclass(frozen=True)
class ParsedAltTextIssue:
    country: str
    url: str
    card_index: int | None
    src: str | None
    srcset: str | None
    alt: str | None
    alt_length: int | None
    alt_comment: str | None


def parse_alt_text_issues(file_path: Path, sheet_name: str = "Feature Card Alt Text Error") -> list[ParsedAltTextIssue]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        header_row = None
        for r in range(1, 10):
            values = [normalize_header(c.value) for c in ws[r]]
            if "COUNTRY" in values and "URL" in values:
                header_row = r
                break
        if header_row is None:
            return []

        header = [normalize_header(c.value) for c in ws[header_row]]
        col = {name: header.index(name) for name in header if name}

        def idx(name: str) -> int | None:
            return col.get(name)

        required = {"COUNTRY", "URL"}
        if not required.issubset(set(col.keys())):
            return []

        issues: list[ParsedAltTextIssue] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            country = clean_cell(row[idx("COUNTRY")] if idx("COUNTRY") is not None else None)
            url = clean_cell(row[idx("URL")] if idx("URL") is not None else None)
            if not country or not url:
                continue

            issues.append(
                ParsedAltTextIssue(
                    country=str(country),
                    url=str(url),
                    card_index=to_int(row[idx("CARDINDEX")] if idx("CARDINDEX") is not None else None),
                    src=str(clean_cell(row[idx("SRC")] if idx("SRC") is not None else None) or "") or None,
                    srcset=str(clean_cell(row[idx("SRCSET")] if idx("SRCSET") is not None else None) or "") or None,
                    alt=str(clean_cell(row[idx("ALT")] if idx("ALT") is not None else None) or "") or None,
                    alt_length=to_int(row[idx("ALTLENGTH")] if idx("ALTLENGTH") is not None else None),
                    alt_comment=str(clean_cell(row[idx("ALTCOMMENT")] if idx("ALTCOMMENT") is not None else None) or "") or None,
                )
            )
        return issues
    finally:
        wb.close()
